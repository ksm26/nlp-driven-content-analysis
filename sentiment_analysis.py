import os
import pandas as pd
import spacy
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load spaCy English tokenizer
nlp = spacy.load("en_core_web_sm")

# Define directories
STOPWORDS_DIR = "StopWords"
DICTIONARY_DIR = "MasterDictionary"
TEXT_FILES_DIR = "extracted_articles"
OUTPUT_EXCEL = "Output Data Structure.xlsx"

# Function to load stop words
def load_stop_words():
    stop_words = set()
    for filename in os.listdir(STOPWORDS_DIR):
        filepath = os.path.join(STOPWORDS_DIR, filename)
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
            stop_words.update(file.read().splitlines())
    return stop_words

# Function to load positive and negative words
def load_sentiment_words():
    with open(os.path.join(DICTIONARY_DIR, "positive-words.txt"), 'r', encoding='utf-8') as file:
        positive_words = set(file.read().splitlines())
    with open(os.path.join(DICTIONARY_DIR, "negative-words.txt"), 'r', encoding='utf-8', errors='ignore') as file:
        negative_words = set(file.read().splitlines())
    return positive_words, negative_words

# Load stop words and sentiment words
stop_words = load_stop_words()
positive_words, negative_words = load_sentiment_words()

# Function to tokenize text using spaCy
def spacy_tokenize(text):
    doc = nlp(text.lower())
    return [token.text for token in doc if token.is_alpha and token.text not in stop_words]

# Function to count complex words
def count_complex_words(words):
    return sum(1 for word in words if sum(1 for char in word if char in "aeiouAEIOU") > 2)

# Function to count syllables in a word
def count_syllables(word):
    word = word.lower()
    vowels = "aeiou"
    syllable_count = sum(1 for char in word if char in vowels)
    if word.endswith("es") or word.endswith("ed"):
        syllable_count -= 1
    return max(syllable_count, 1)

# Function to count personal pronouns
def count_personal_pronouns(text):
    pronouns = re.findall(r'\b(I|we|my|ours|us)\b', text, re.I)
    return len(pronouns)

# Function to calculate readability metrics
def calculate_readability(text):
    sentences = list(nlp(text).sents)
    words = spacy_tokenize(text)
    num_words = len(words)
    num_sentences = len(sentences)
    complex_word_count = count_complex_words(words)
    
    avg_sentence_length = num_words / num_sentences if num_sentences else 0
    percent_complex_words = complex_word_count / num_words if num_words else 0
    fog_index = 0.4 * (avg_sentence_length + percent_complex_words)
    avg_words_per_sentence = num_words / num_sentences if num_sentences else 0
    
    syllable_count_per_word = sum(count_syllables(word) for word in words) / num_words if num_words else 0
    personal_pronouns = count_personal_pronouns(text)
    avg_word_length = sum(len(word) for word in words) / num_words if num_words else 0
    
    return avg_sentence_length, percent_complex_words, fog_index, avg_words_per_sentence, complex_word_count, num_words, syllable_count_per_word, personal_pronouns, avg_word_length

# Function to calculate sentiment scores
def calculate_sentiment_scores(text):
    words = spacy_tokenize(text)
    
    positive_score = sum(1 for word in words if word in positive_words)
    negative_score = sum(1 for word in words if word in negative_words)
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    subjectivity_score = (positive_score + negative_score) / (len(words) + 0.000001)
    
    return positive_score, negative_score, polarity_score, subjectivity_score

# Process each text file and compute metrics
results = []
for filename in os.listdir(TEXT_FILES_DIR):
    if filename.endswith(".txt"):
        filepath = os.path.join(TEXT_FILES_DIR, filename)
        with open(filepath, 'r', encoding='utf-8') as file:
            text = file.read()
        
        pos_score, neg_score, pol_score, subj_score = calculate_sentiment_scores(text)
        avg_sent_len, pct_complex, fog_idx, avg_words_sent, complex_count, word_count, syllables_word, pronouns, avg_word_len = calculate_readability(text)
        
        results.append([filename[:-4], pos_score, neg_score, pol_score, subj_score, avg_sent_len, pct_complex, fog_idx, avg_words_sent, complex_count, word_count, syllables_word, pronouns, avg_word_len])

# Load existing Excel file and update columns
df_excel = pd.read_excel(OUTPUT_EXCEL)
df_results = pd.DataFrame(results, columns=["URL_ID", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE", "SUBJECTIVITY SCORE", "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS", "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE", "COMPLEX WORD COUNT", "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"])

df_merged = df_excel.merge(df_results, on="URL_ID", how="left")

# Drop old columns (_x) and rename new columns (_y)
for col in df_merged.columns:
    if col.endswith("_x"):  
        df_merged.drop(columns=[col], inplace=True)  # Remove old duplicate columns
    elif col.endswith("_y"):
        df_merged.rename(columns={col: col[:-2]}, inplace=True)  # Remove _y suffix

ordered_columns = ["URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE", "SUBJECTIVITY SCORE", "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS", "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE", "COMPLEX WORD COUNT", "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"]

df_merged = df_merged[ordered_columns]
df_merged.to_excel(OUTPUT_EXCEL, index=False)

# Adjust column widths
wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active
ws.column_dimensions[get_column_letter(1)].width = 25  # Set first column width larger
for col in range(2, len(ordered_columns) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 15
wb.save(OUTPUT_EXCEL)

print(f"Analysis completed. Updated results saved in {OUTPUT_EXCEL}")
